# core/report_generator.py
# Comprehensive Excel Report Generator with Charts and Analysis

import io
from datetime import datetime, date
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class FinancialReportGenerator:
    """
    Generate comprehensive financial reports in Excel format.
    Includes multiple sheets with data, charts, and analysis.
    """
    
    def __init__(self):
        # Define color scheme
        self.colors = {
            'header': 'FF4472C4',
            'subheader': 'FF5B9BD5',
            'income': 'FF70AD47',
            'expense': 'FFC5504',
            'savings': 'FFFFC000',
            'neutral': 'FFD9E1F2'
        }
        
        # Define border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_comprehensive_report(
        self,
        user_data: Dict,
        transaction_summary: Dict,
        categories: List[Dict],
        monthly_trend: Optional[pd.DataFrame] = None,
        goals: Optional[List[Dict]] = None,
        investment_options: Optional[Dict] = None
    ) -> io.BytesIO:
        """
        Generate comprehensive Excel report.
        
        Args:
            user_data: User profile information
            transaction_summary: Financial summary data
            categories: Category-wise breakdown
            monthly_trend: Monthly trend data
            goals: Financial goals
            investment_options: Investment recommendations
            
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, user_data, transaction_summary)
        self._create_category_breakdown_sheet(wb, categories)
        
        if monthly_trend is not None and not monthly_trend.empty:
            self._create_trend_analysis_sheet(wb, monthly_trend)
        
        self._create_tax_optimization_sheet(wb, transaction_summary)
        
        if goals:
            self._create_goals_tracker_sheet(wb, goals)
        
        if investment_options:
            self._create_investment_options_sheet(wb, investment_options)
        
        self._create_recommendations_sheet(wb, transaction_summary, categories)
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _create_summary_sheet(self, wb: Workbook, user_data: Dict, summary: Dict):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "Financial Analysis Report"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Report metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A3'] = f"User: {user_data.get('username', 'N/A')}"
        ws['A4'] = f"Period: {summary.get('date_range', 'N/A')}"
        
        # Financial Overview
        ws['A6'] = "Financial Overview"
        ws['A6'].font = Font(size=14, bold=True)
        
        headers = ['Metric', 'Amount (₹)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        income = summary.get('total_income', 0)
        expenses = summary.get('total_expenses', 0)
        savings = summary.get('savings', 0)
        savings_rate = summary.get('savings_rate', 0)
        
        data = [
            ['Total Income', income, '✓'],
            ['Total Expenses', expenses, '✓'],
            ['Net Savings', savings, '✓' if savings > 0 else '✗'],
            ['Savings Rate', f"{savings_rate:.1f}%", self._get_savings_status(savings_rate)]
        ]
        
        for idx, row_data in enumerate(data, 8):
            ws.cell(row=idx, column=1, value=row_data[0])
            ws.cell(row=idx, column=2, value=row_data[1] if isinstance(row_data[1], str) else f"₹{row_data[1]:,.2f}")
            ws.cell(row=idx, column=3, value=row_data[2])
            
            # Color code based on type
            if 'Income' in row_data[0]:
                ws.cell(row=idx, column=2).fill = PatternFill(start_color=self.colors['income'], end_color=self.colors['income'], fill_type='solid')
            elif 'Expenses' in row_data[0]:
                ws.cell(row=idx, column=2).fill = PatternFill(start_color=self.colors['expense'], end_color=self.colors['expense'], fill_type='solid')
            elif 'Savings' in row_data[0]:
                ws.cell(row=idx, column=2).fill = PatternFill(start_color=self.colors['savings'], end_color=self.colors['savings'], fill_type='solid')
        
        # Key Insights
        ws['A13'] = "Key Insights"
        ws['A13'].font = Font(size=12, bold=True)
        
        insights = self._generate_insights(summary, income, expenses, savings, savings_rate)
        for idx, insight in enumerate(insights, 14):
            ws.cell(row=idx, column=1, value=f"• {insight}")
            ws.merge_cells(f'A{idx}:E{idx}')
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_category_breakdown_sheet(self, wb: Workbook, categories: List[Dict]):
        """Create category breakdown sheet with chart."""
        ws = wb.create_sheet("Category Breakdown")
        
        # Title
        ws['A1'] = "Spending by Category"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Category', 'Amount (₹)', 'Percentage', 'Transaction Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Calculate total for percentages
        total = sum(cat.get('amount', 0) for cat in categories)
        
        # Data rows
        for idx, category in enumerate(categories, 4):
            amount = category.get('amount', 0)
            percentage = (amount / total * 100) if total > 0 else 0
            
            ws.cell(row=idx, column=1, value=category.get('category', 'Unknown').replace('_', ' ').title())
            ws.cell(row=idx, column=2, value=amount)
            ws.cell(row=idx, column=2).number_format = '₹#,##0.00'
            ws.cell(row=idx, column=3, value=percentage / 100)
            ws.cell(row=idx, column=3).number_format = '0.00%'
            ws.cell(row=idx, column=4, value=category.get('transaction_count', 0))
        
        # Add pie chart
        if categories:
            pie_chart = PieChart()
            labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(categories))
            data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(categories))
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            pie_chart.title = "Spending Distribution"
            ws.add_chart(pie_chart, "F3")
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
    
    def _create_trend_analysis_sheet(self, wb: Workbook, monthly_trend: pd.DataFrame):
        """Create monthly trend analysis sheet with line chart."""
        ws = wb.create_sheet("Trend Analysis")
        
        # Title
        ws['A1'] = "Monthly Trend Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Convert dataframe to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(monthly_trend, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:  # Header row
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Add line chart
        if len(monthly_trend) > 1:
            line_chart = LineChart()
            line_chart.title = "Income vs Expenses Over Time"
            line_chart.y_axis.title = "Amount (₹)"
            line_chart.x_axis.title = "Month"
            
            data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=3 + len(monthly_trend))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(monthly_trend))
            
            line_chart.add_data(data, titles_from_data=True)
            line_chart.set_categories(cats)
            
            ws.add_chart(line_chart, "E3")
        
        # Format columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_tax_optimization_sheet(self, wb: Workbook, summary: Dict):
        """Create tax optimization recommendations sheet."""
        ws = wb.create_sheet("Tax Optimization")
        
        # Title
        ws['A1'] = "Tax Saving Opportunities"
        ws['A1'].font = Font(size=14, bold=True)
        
        income = summary.get('total_income', 0)
        
        # Section 80C
        ws['A3'] = "Section 80C (up to ₹1.5 Lakh)"
        ws['A3'].font = Font(size=12, bold=True)
        
        section_80c = [
            ['ELSS Mutual Funds', '₹1,50,000', 'High growth potential, 3-year lock-in'],
            ['PPF', '₹1,50,000', 'Safe, 15-year lock-in, tax-free returns'],
            ['Tax Saver FD', '₹1,50,000', 'Safe, 5-year lock-in, fixed returns'],
            ['Life Insurance Premium', '₹1,50,000', 'Protection + tax benefit'],
            ['NSC', '₹1,50,000', 'Safe, 5-year lock-in']
        ]
        
        headers = ['Investment Option', 'Max Deduction', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        for idx, row_data in enumerate(section_80c, 5):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=idx, column=col, value=value)
        
        # Additional sections
        ws['A11'] = "Additional Tax Saving Sections"
        ws['A11'].font = Font(size=12, bold=True)
        
        additional_sections = [
            ['80CCD(1B)', '₹50,000', 'NPS contribution'],
            ['80D', '₹25,000-₹100,000', 'Health insurance premium'],
            ['80E', 'No limit', 'Education loan interest'],
            ['24(b)', '₹2,00,000', 'Home loan interest']
        ]
        
        for col, header in enumerate(['Section', 'Max Deduction', 'Description'], 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        for idx, row_data in enumerate(additional_sections, 13):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=idx, column=col, value=value)
        
        # Estimated tax savings
        ws['A18'] = "💡 Potential Tax Savings"
        ws['A18'].font = Font(size=12, bold=True)
        
        max_savings_80c = 150000
        max_savings_80ccd = 50000
        total_deduction = max_savings_80c + max_savings_80ccd
        
        # Assume 30% tax bracket
        tax_saved = total_deduction * 0.30
        
        ws['A19'] = f"If you maximize deductions (₹{total_deduction:,.0f}):"
        ws['A20'] = f"Estimated Tax Savings: ₹{tax_saved:,.0f}"
        ws['A20'].font = Font(bold=True, size=11)
        ws['A20'].fill = PatternFill(start_color=self.colors['savings'], end_color=self.colors['savings'], fill_type='solid')
        
        # Format columns
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_goals_tracker_sheet(self, wb: Workbook, goals: List[Dict]):
        """Create goals tracking sheet."""
        ws = wb.create_sheet("Goals Tracker")
        
        # Title
        ws['A1'] = "Financial Goals Tracker"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Goal Name', 'Target Amount', 'Current Savings', 'Progress %', 'Monthly SIP', 'Due Date', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Data rows
        for idx, goal in enumerate(goals, 4):
            target = goal.get('target_amount', 0)
            current = goal.get('current_balance', 0)
            progress = (current / target * 100) if target > 0 else 0
            
            ws.cell(row=idx, column=1, value=goal.get('goal_name', 'Unnamed Goal'))
            ws.cell(row=idx, column=2, value=target)
            ws.cell(row=idx, column=2).number_format = '₹#,##0.00'
            ws.cell(row=idx, column=3, value=current)
            ws.cell(row=idx, column=3).number_format = '₹#,##0.00'
            ws.cell(row=idx, column=4, value=progress / 100)
            ws.cell(row=idx, column=4).number_format = '0.00%'
            ws.cell(row=idx, column=5, value=goal.get('monthly_contribution', 0))
            ws.cell(row=idx, column=5).number_format = '₹#,##0.00'
            ws.cell(row=idx, column=6, value=goal.get('due_date', 'N/A'))
            
            # Status
            if progress >= 100:
                status = '✓ Complete'
            elif progress >= 75:
                status = '→ On Track'
            elif progress >= 50:
                status = '⚠ Attention Needed'
            else:
                status = '⚠ Behind Schedule'
            
            ws.cell(row=idx, column=7, value=status)
        
        # Format columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_investment_options_sheet(self, wb: Workbook, investment_options: Dict):
        """Create investment options recommendations sheet."""
        ws = wb.create_sheet("Investment Options")
        
        # Title
        ws['A1'] = "Investment Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A2'] = f"Goal: {investment_options.get('goal_name', 'N/A')}"
        ws['A3'] = f"Time Horizon: {investment_options.get('time_horizon_months', 0)} months"
        ws['A4'] = f"Expected Return: {investment_options.get('overall_expected_return', 0) * 100:.2f}%"
        
        # Headers
        headers = ['Investment Vehicle', 'Allocation %', 'Expected Return', 'Risk Level', 'Liquidity (Days)', 'Min Investment']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Data rows
        options = investment_options.get('recommended_options', [])
        for idx, option in enumerate(options, 7):
            ws.cell(row=idx, column=1, value=option.get('name', ''))
            ws.cell(row=idx, column=2, value=option.get('allocation_percentage', 0) / 100)
            ws.cell(row=idx, column=2).number_format = '0.00%'
            ws.cell(row=idx, column=3, value=option.get('expected_return_annual', 0) / 100)
            ws.cell(row=idx, column=3).number_format = '0.00%'
            ws.cell(row=idx, column=4, value=option.get('risk_category', '').replace('_', ' ').title())
            ws.cell(row=idx, column=5, value=option.get('liquidity_days', 0))
            ws.cell(row=idx, column=6, value=option.get('min_investment', 0))
            ws.cell(row=idx, column=6).number_format = '₹#,##0.00'
        
        # Format columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_recommendations_sheet(self, wb: Workbook, summary: Dict, categories: List[Dict]):
        """Create personalized recommendations sheet."""
        ws = wb.create_sheet("Recommendations")
        
        # Title
        ws['A1'] = "Personalized Financial Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(summary, categories)
        
        row = 3
        for category, recs in recommendations.items():
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=1).font = Font(size=12, bold=True)
            row += 1
            
            for rec in recs:
                ws.cell(row=row, column=1, value=f"• {rec}")
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
            
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 80
    
    def _get_savings_status(self, savings_rate: float) -> str:
        """Get savings status emoji."""
        if savings_rate >= 30:
            return '🌟 Excellent'
        elif savings_rate >= 20:
            return '✓ Good'
        elif savings_rate >= 10:
            return '→ Fair'
        else:
            return '⚠ Needs Improvement'
    
    def _generate_insights(self, summary: Dict, income: float, expenses: float, savings: float, savings_rate: float) -> List[str]:
        """Generate key insights."""
        insights = []
        
        if savings_rate >= 30:
            insights.append(f"Excellent! Your savings rate of {savings_rate:.1f}% is well above the recommended 20%")
        elif savings_rate >= 20:
            insights.append(f"Good job! You're saving {savings_rate:.1f}% of your income")
        else:
            insights.append(f"Consider increasing your savings rate from {savings_rate:.1f}% to at least 20%")
        
        if income > 0:
            expense_ratio = (expenses / income) * 100
            if expense_ratio > 80:
                insights.append(f"Your expense ratio is {expense_ratio:.1f}%. Consider reducing discretionary spending")
        
        txn_count = summary.get('transaction_count', 0)
        insights.append(f"Analyzed {txn_count} transactions across {summary.get('date_range', 'the period')}")
        
        return insights
    
    def _generate_recommendations(self, summary: Dict, categories: List[Dict]) -> Dict[str, List[str]]:
        """Generate personalized recommendations."""
        recs = {}
        
        # Savings recommendations
        savings_rate = summary.get('savings_rate', 0)
        if savings_rate < 20:
            recs['Increase Savings'] = [
                'Aim to save at least 20% of your income',
                'Set up automatic transfers to savings account on payday',
                'Use the 50/30/20 rule: 50% needs, 30% wants, 20% savings'
            ]
        
        # Category-based recommendations
        if categories:
            top_category = max(categories, key=lambda x: x.get('amount', 0))
            top_amount = top_category.get('amount', 0)
            total_expenses = sum(cat.get('amount', 0) for cat in categories)
            
            if total_expenses > 0:
                top_percentage = (top_amount / total_expenses) * 100
                if top_percentage > 30:
                    category_name = top_category.get('category', 'Unknown').replace('_', ' ').title()
                    recs['Expense Optimization'] = [
                        f'{category_name} accounts for {top_percentage:.1f}% of your spending',
                        f'Consider setting a monthly budget limit for {category_name}',
                        'Track and review this category weekly'
                    ]
        
        # Tax optimization
        recs['Tax Optimization'] = [
            'Maximize Section 80C deductions (up to ₹1.5 lakh)',
            'Consider NPS contribution for additional ₹50,000 deduction under 80CCD(1B)',
            'Explore health insurance for 80D benefits'
        ]
        
        # Investment recommendations
        recs['Investment Strategy'] = [
            'Maintain 3-6 months expenses as emergency fund',
            'Start SIPs in diversified mutual funds for long-term goals',
            'Review and rebalance portfolio annually'
        ]
        
        return recs